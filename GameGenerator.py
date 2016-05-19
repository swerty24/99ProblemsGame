# -*- coding: utf-8 -*-
"""
5-5-2016 by: Swerty
This is a open source python script to generate the 99 Problems Game as an excel spreadsheet
The end result should be able to (1) generate a spreadsheet of all the problems and
solutions as a printable excel file. It should also be able to (2) add and remove problems
or solutions to the complete list through a text command line.
--------------------
This script requires Openpyxl package:http://openpyxl.readthedocs.io/en/default/
--------------------
TODO:
Modify for length of problems for both Row's and Columns.
Determine the length of problems for a 4 column Sheet.
Determine max characters at each font size. To auto-scale text size in cell.
Output a problems and solutions sheet. With RED on the back of the problems card for printing on both sides.
Make Squares bigger
QR codes
Print sheet and see how big each game piece is

Part 2 of requirements
Create menu for interaction with problem and solution list. Functions include:
-list, add, count, or remove problems or solutions
-Generate xlsx (or pdf) of game for printing
seperate dict lists into sub lists. (ie. base, exploritory, sci-fi, silly, etc.)
    and allow for combining of some lists if wanted and igrnored others.
"""

#Imports
import math, string
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter, Cell
from openpyxl.styles import Fill, Color, colors, PatternFill, Border, Side, Alignment, Font
from string import ascii_uppercase




# function to count the number of characters in a string
def LetterCount(str):
            str= str.lower().strip()
            str = str.strip(string.punctuation)
            list1=list(str)
            count= 0
            for l in list1:
                    if l.isalpha():
                        count +=1   
            return count
            
            #Set border thickness for styling cells
medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

#List of Problems (currently just a sample for testing)
problems = {
            1:'World Hunger',
            2:'World Peace',
            3:'Traffic Congestion',
            4:'Air Pollution',
            5:'Global Warming',
            6:'Clean Drinking Water',
            7:'Cancer',
            8:'Clogged Toilets',
            9:'Looking down at a smartphone',
            10:'Desalination',
            11:'Dependence on Foreign Oil',
            12:'Homelessness',
            13:'Terrorism',
            14:'AIDS',
            15:'Sleeping in',
            16:'Smelly Socks',
            17:'Elephant Tusk Poaching',
            18:'Obesity',
            19:'Rapid Inflation',
            20:'School Shootings',
            21:'Too Many Lawyers',
            22:'Human Trafficking',
            23:'Noise Pollution',
            24:'Alzheimers',
            25:'Illegal Drug Trade',
            26:'Drunk Driving',
            27:'Hangovers',
            28:'Coral Reef Destruction',
            29:'Manatees being chopped up by boat propoellers',
            30:'Rising college costs',
            31:'Over Population',
            32:'Political Bi-partisanship in the USA',
            33:'Genocide',
            34:'Homicide',
            35:'Suicide',
            36:'Rising Healthcare Costs',
            37:'Rising Prison Costs',
            38:'Unemployment',
            39:'China',
            40:'Job Outsourcing',
            41:'Birds Getting chopped by Wind turbines',
            42:'Forest Fires',
            43:'Gender Disparity in the Workforce',
            44:'Gang Violence',
            45:'Deforestation',
            46:'Sudden Infant Death Syndrome',
            47:'Texting while Driving',
            48:'Insider Trading',
            49:'Binge Watching TV Shows',
            50:'Car Accidents',
            51:'Not having enough time in the day',
            52:'Shortage of Teachers',
            53:'Shark Attacks',
            54:'Donald Trump',
            55:'Waste Disposal',
            56:'Loss of Biodiversity ',
            57:'Ocean Acidification ',
            58:'Choosing the wrong check out lane',
            59:'Power Outages ',
            60:'Peak Oil ',
            61:'Apathy ',
            62:'Earthquakes '
            }
            
solutions = {
            1:'A Plunger',
            2:'Google Glass',
            3:'Self Driving Cars',
            4:'Teleportation',
            5:'Gravity',
            6:'Perpetual Motion Machine',
            7:'Hydroponics',
            8:'Reverse Osmosis',
            9:'Mobile Phones',
            10:'Airfoil',
            11:'Global Positioning System',
            12:'Google Search',
            13:'The Cathode Ray Tube',
            14:'Drones',
            15:'Triangulation',
            16:'The GoPro Camera',
            17:'Solar Panels',
            18:'The Inkjet Printer',
            19:'The Greenhouse Effect',
            20:'Cards Against Humanity',
            21:'A Rocket',
            22:'Morse Code',
            23:'The Ujjayi Breath',
            24:'Social Networks',
            25:'Integration By Parts',
            26:'Photosynthesis',
            27:'The Coriolis Effect',
            28:'The Haber-Bosh Process',
            29:'Resonance',
            30:'Vegetarianism',
            31:'USB Ports',
            32:'Neodymium Magnets',
            33:'The Bubble Sort Algorithm',
            34:'Latex',
            35:'Plastic',
            36:'Telemetry',
            37:'Worldwide Satellite Internet',
            38:'Electric Cars',
            39:'Fire',
            40:'3D Printing',
            41:'Sporks',
            42:'Chocolate',
            43:'Clay',
            44:'Gorilla Glue',
            45:'Peanut Butter',
            46:'Brownian Motion',
            47:'Neural Network Programming',
            48:'Genetic Algorithm',
            49:'The God Particle',
            50:'Online Education',
            51:'Adaptation by Natural Selection',
            52:'Cargo Bikes',
            53:'Linux',
            54:'Open Source',
            55:'Magma',
            56:'C++',
            57:'Unmanned Solar Powered Aircraft',
            58:'Weather Balloons',
            59:'Friendship',
            60:'Touch Screens',
            61:'Community Gardens',
            62:'Free Wifi',
            63:'Pocket Parks',
            64:'The Printed Newspaper',
            65:'Snail Mail',
            66:'Bioluminescence',
            67:'Van Der Waals Force',
            68:'Oculus Rift',
            69:'Craigslist',
            70:'Becoming a Multiplanetary Species',
            71:'The Transistor',
            72:'Hyperloop',
            73:'Biomimetics',
            74:'Donald Trump',
            75:'Drinking Fountains ',
            76:'Phone Charging Stations ',
            77:'The Keurig ',
            78:'Chuck Norris ',
            79:'Composting',
            80:'Iron Fertilixation Effect'
            }


#Create Workbook using open Pyxl package
wb = Workbook()
# Set Workbook File Name and Sheet Name
dest_filename = '99ProblemsGame - Printable.xlsx'
ws1 = wb.active
ws1.title = "To Print"

           

#Set the Color of Cell
rgb=[255,0,0]
#Set the Color of string
color_string="0000"
#color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])

#Calculate the number of rows needed (based on 4 columns) and store into x
x = int(math.ceil(len(problems)/4.0))+1
print x                
#Loop through all the columns upto D, and set each cell to red and set size of text for cell
for ProblemColumnLetter in ascii_uppercase:
    #Only use 4 Columns with width 30 
    ws1.column_dimensions[ProblemColumnLetter].width = 30
    for i in range(1,x):
        ws1[ProblemColumnLetter+str(i)].fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
        #Set text size based on number of characters in dict list and Color to White
        if LetterCount(problems[i]) < 10:
            ws1[ProblemColumnLetter+str(i)].font = Font(size = 19, color=colors.WHITE)
        else:
            ws1[ProblemColumnLetter+str(i)].font = Font(size = 23, color=colors.WHITE)
    
        print ProblemColumnLetter + str(i)
    if ProblemColumnLetter == "D":
        break
       


#print each of the problems into the cells
i = 1
for row in range(1, x):
    for col in range(1, 5):
            _ = ws1.cell(column=col, row=row,value="%s" % problems[i])
                      
            if i< len(problems):
                i+= 1
            else:     
                break
            
#loop through all the cells and apply height and border      
i = 1
for row in range(1, int(math.ceil(len(problems)/4.0))+1):
    #set row height 
    ws1.row_dimensions[row].height = 55
    for col in range(1, 5):
            #apply border
            ws1.cell(row=row, column=col).border=medium_border
           
                      
wb.save(filename = dest_filename)      



            


#The following is example code for working with openpyxl            
"""               
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
print(problems[2])
"""
